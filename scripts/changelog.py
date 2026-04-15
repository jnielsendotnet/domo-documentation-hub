#!/usr/bin/env python3
"""
Generate a changelog spreadsheet from git history for the last N days.

Outputs an Excel file (or CSV with --csv) listing every MDX file changed in
the window, with article title, localized titles, commit-message summaries,
first-created date, and last-updated date.

Usage:
    python scripts/changelog.py [--days=7] [--csv] [--output=<path>]

Requirements (Excel output only):
    pip install openpyxl

Run it:

# Default — last 7 days, Excel output in scripts/reports/
python3 scripts/changelog.py

# Custom window
python3 scripts/changelog.py --days=14

# CSV instead of Excel
python3 scripts/changelog.py --csv

# Custom output path
python scripts/changelog.py --output=~/Desktop/changelog.xlsx
Output lands in scripts/reports/ as changelog_YYYYMMDD_HHMMSS.xlsx (or .csv).

Columns in the spreadsheet:

Column	Source
File Name	Git path relative to repo root
Article Name	title from MDX frontmatter
DE / ES / FR / JA Article Name	title from matching locale file under de/, es/, etc.
Changes Made	Semicolon-separated commit subjects that touched the file
First Created Date	Oldest commit in full git history (follows renames)
Last Updated Date	Most recent commit


"""

import csv
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
LOCALE_DIRS = ["de", "es", "fr", "ja"]
DEFAULT_DAYS = 7
REPORTS_DIR = Path(__file__).resolve().parent / "reports"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Git helpers
# ---------------------------------------------------------------------------

def git(*args: str) -> str:
    result = subprocess.run(
        ["git", *args],
        capture_output=True,
        text=True,
        cwd=REPO_ROOT,
    )
    return result.stdout.strip()


def get_changed_mdx_files(since_days: int) -> list[str]:
    """Return sorted list of MDX paths (relative to repo root) changed in the window."""
    raw = git(
        "log",
        f"--since={since_days} days ago",
        "--name-only",
        "--pretty=format:",
        "--diff-filter=ACDMRT",
    )
    seen: set[str] = set()
    for line in raw.splitlines():
        path = line.strip()
        if path.endswith(".mdx"):
            seen.add(path)
    return sorted(seen)


def get_commit_messages(rel_path: str, since_days: int) -> str:
    """Return semicolon-separated commit subjects for this file in the window."""
    raw = git(
        "log",
        f"--since={since_days} days ago",
        "--format=%s",
        "--",
        rel_path,
    )
    messages = [m.strip() for m in raw.splitlines() if m.strip()]
    return "; ".join(messages)


def get_file_dates(rel_path: str) -> tuple[str, str]:
    """Return (first_created, last_updated) as formatted date strings."""
    last_raw = git("log", "-1", "--format=%ai", "--", rel_path)
    all_raw = git("log", "--follow", "--format=%ai", "--", rel_path)

    def fmt(iso: str) -> str:
        iso = iso.strip()
        if not iso:
            return ""
        try:
            # git %ai: "2026-04-14 15:42:18 -0700"
            dt = datetime.strptime(iso[:19], "%Y-%m-%d %H:%M:%S")
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            return iso

    first = fmt(all_raw.splitlines()[-1]) if all_raw else ""
    last = fmt(last_raw)
    return first, last


# ---------------------------------------------------------------------------
# Content helpers
# ---------------------------------------------------------------------------

def extract_title(file_path: Path) -> str:
    """Read YAML frontmatter title from an MDX file."""
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except FileNotFoundError:
        return ""
    match = re.search(
        r'^---\s*\n.*?^title:\s*["\']?(.+?)["\']?\s*$',
        text,
        re.MULTILINE | re.DOTALL,
    )
    if match:
        return match.group(1).strip().strip("\"'")
    return ""


def get_localized_titles(rel_path: str) -> dict[str, str]:
    """Return {locale: title} for each locale that has a matching file."""
    titles: dict[str, str] = {}
    for locale in LOCALE_DIRS:
        localized = REPO_ROOT / locale / rel_path
        title = extract_title(localized)
        if title:
            titles[locale] = title
    return titles


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def build_rows(since_days: int) -> list[dict]:
    files = get_changed_mdx_files(since_days)
    print(f"Found {len(files)} changed MDX file(s) in the last {since_days} day(s).")

    rows = []
    for i, rel_path in enumerate(files, 1):
        print(f"  [{i}/{len(files)}] {rel_path}")
        abs_path = REPO_ROOT / rel_path
        article_name = extract_title(abs_path) if abs_path.exists() else ""
        localized = get_localized_titles(rel_path)
        changes = get_commit_messages(rel_path, since_days)
        first_created, last_updated = get_file_dates(rel_path)

        rows.append({
            "File Name": rel_path,
            "Article Name": article_name,
            "DE Article Name": localized.get("de", ""),
            "ES Article Name": localized.get("es", ""),
            "FR Article Name": localized.get("fr", ""),
            "JA Article Name": localized.get("ja", ""),
            "Changes Made": changes,
            "First Created Date": first_created,
            "Last Updated Date": last_updated,
        })
    return rows


# ---------------------------------------------------------------------------
# Output: CSV
# ---------------------------------------------------------------------------

COLUMNS = [
    "File Name",
    "Article Name",
    "DE Article Name",
    "ES Article Name",
    "FR Article Name",
    "JA Article Name",
    "Changes Made",
    "First Created Date",
    "Last Updated Date",
]


def write_csv(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Saved CSV → {out_path}")


# ---------------------------------------------------------------------------
# Output: Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1A3254") if HAS_OPENPYXL else None  # Domo navy
HEADER_FONT = Font(bold=True, color="FFFFFF") if HAS_OPENPYXL else None

COL_WIDTHS = {
    "File Name": 50,
    "Article Name": 40,
    "DE Article Name": 36,
    "ES Article Name": 36,
    "FR Article Name": 36,
    "JA Article Name": 36,
    "Changes Made": 60,
    "First Created Date": 20,
    "Last Updated Date": 20,
}


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Changelog"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(col_name, 20)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    wrap = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.alignment = wrap

    wb.save(out_path)
    print(f"Saved Excel → {out_path}")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def parse_args() -> tuple[int, bool, Path | None]:
    days = DEFAULT_DAYS
    use_csv = "--csv" in sys.argv
    output: Path | None = None

    for arg in sys.argv[1:]:
        if arg.startswith("--days="):
            days = int(arg.split("=", 1)[1])
        elif arg == "--csv":
            pass  # handled above
        elif arg.startswith("--output="):
            output = Path(arg.split("=", 1)[1])

    return days, use_csv, output


def main() -> None:
    days, use_csv, output = parse_args()

    print(f"Scanning git history for the last {days} day(s)…")
    rows = build_rows(days)

    if not rows:
        print("No MDX files changed in that window. Nothing to write.")
        sys.exit(0)

    date_str = datetime.now().strftime("%Y-%m-%d")

    if use_csv or not HAS_OPENPYXL:
        if not HAS_OPENPYXL and not use_csv:
            print("openpyxl not installed — falling back to CSV. Run: pip install openpyxl")
        out_path = output or REPORTS_DIR / f"changelog-{date_str}.csv"
        write_csv(rows, out_path)
    else:
        out_path = output or REPORTS_DIR / f"changelog-{date_str}.xlsx"
        write_xlsx(rows, out_path)


if __name__ == "__main__":
    main()
